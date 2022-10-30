import os

import pdfplumber
import xlwt
import xlrd

from dto.material import Material
from openpyxl import load_workbook
from openpyxl.comments import Comment

fromPDFRootFolderPath = r"C:\Users\moosk\PycharmProjects\pythonProject\Office2PDF\test\sub\pdf"
fromExcelRootFolderPath = r"C:\Users\moosk\PycharmProjects\pythonProject\Office2PDF\test\sub\Excel"
outPutFolderPath = r"C:\Users\moosk\PycharmProjects\pythonProject\Office2PDF\test\sub\Excel"

manul = [
    {"铸锭批号": "J5306-12341",
     "时间": "2022/10/18"
     },
    {"铸锭批号": "J5203-12278",
     "时间": "2022/10/08"
     }
]


def pdf_to_excel():
    for root, dirs, files in os.walk(fromPDFRootFolderPath):
        for file in files:
            print(f"root is {root}")
            file_name = str.removesuffix(file, ".pdf")
            path = os.path.join(fromPDFRootFolderPath, file)
            workbook = xlwt.Workbook()  # 定义workbook
            sheet = workbook.add_sheet('Sheet1')  # 添加sheet
            print(path)
            i = 0  # Excel起始位置
            pdf = pdfplumber.open(path)
            print('\n')
            print('开始读取数据')
            print('\n')
            for page in pdf.pages:
                # 获取当前页面的全部文本信息，包括表格中的文字
                # print(page.extract_text())
                for table in page.extract_tables():
                    # print(table)
                    for row in table:
                        print(row)
                        for j in range(len(row)):
                            sheet.write(i, j, row[j])
                        i += 1
                    print('---------- 分割线 ----------')

            pdf.close()

            # 保存Excel表
            output_path = os.path.join(outPutFolderPath, file_name)
            output_file = f'{output_path}.xls'
            print("output_file", output_file)
            workbook.save(output_file)
            print('\n')
            print('写入excel成功')
    return


def excel_2_convert():
    material_convert = []
    for root, dirs, files in os.walk(fromExcelRootFolderPath):
        manul_index = 0
        for file in files:
            print(f"file{file}")
            find_map = {"牌号": "", "规格/mm": "", "任务单号": ""}
            file_name = str.removesuffix(file, ".xls")
            path = os.path.join(fromExcelRootFolderPath, file)
            print(f"path {path}")
            with xlrd.open_workbook(path) as f:
                table = f.sheets()[0]
                for i in range(0, 3):
                    for j in range(0, table.ncols):
                        key = table.cell_value(i, j).replace('\n', '')
                        if key in find_map:
                            for k in range(j + 1, table.ncols):
                                value = table.cell_value(i, k)
                                if value != '':
                                    find_map[key] = value
                                    break
                # 构造convert
                for i in range(6, 16):

                    value_map = {}
                    for j in range(1, table.ncols):  # 从厂家开始
                        value = table.cell_value(i, j).replace('\n', '')

                        if value:
                            value_map[table.cell_value(5, j).replace('\n', '')] = value
                    material = Material.map_2_convert(value_map=value_map)
                    material.brand_id = find_map.get("牌号")
                    material.standard = find_map.get("规格/mm")
                    material.task_id = find_map.get("任务单号")
                    material.time = manul[manul_index].get("时间")
                    material.zhuding_flow_id = manul[manul_index].get("铸锭批号")

                    if material.flow_id == None or material.flow_id == "":
                        continue
                    material_convert.append(material)

                    # print(value_map)
                # 做一些特殊处理
                for i in range(0, len(material_convert)):
                    material = material_convert[i]
                    if material.flow_id != "" and i != 0 and material.material_name == None:
                        material.material_name = material_convert[i - 1].material_name
                        material.material_brand_id = material_convert[i - 1].material_brand_id
            manul_index = manul_index + 1

    return material_convert


def convert_2_result(material_convert):
    Brand = 1
    Factory_name = 2
    Flow_id = 3
    Consume_num = 8

    Pai_hao = 1
    Ren_wu_dan_hao = 2
    Zhu_ding_pi_hao = 3
    Gui_ge = 4
    Time_col = 34
    Tou_liao_list = ["海绵钛", "Al99.6", "AlV55", "AlV85", "AlFe60", "氧化钛", "钼粉", "TiSn80", "AlMo60", "AlSi10",
                     "AlSi12", "铁丝", "HZr-1", "AlNb75", "金属铬", "Cu-30Ti", "DJMnD", "AlCr70", "AlMoCrFeSi",
                     "Ti-32Fe"]
    wb = load_workbook(r"C:\Users\moosk\PycharmProjects\pythonProject\Office2PDF\test\sub\template\汇总表.xlsx")
    sh = wb["原料库存表"]  # 根据表单名称，选择sheet（表单）
    sh_2 = wb["生产明细表"]  # 根据表单名称，选择sheet（表单）
    index = 4
    start_chengchanmingxi = 4
    wirte_map = {}
    list_special = ["海绵钛"]
    comment_author = "xingxuewei"
    for i in range(4, len(material_convert) + 4):
        material = material_convert[i - 6]
        if material.flow_id == "" or material.flow_id == None:
            continue

        # 如果flow_id相同，从8列开始向后找不为空的值写入
        if material.flow_id in wirte_map.keys():
            row = wirte_map[material.flow_id]
            for i in range(Consume_num, Consume_num + 12):
                if sh.cell(row, i).value == None or sh.cell(row, i).value == "":
                    sh.cell(row, i).value = material.consume_num
                    comment = Comment(material.zhuding_flow_id, comment_author)
                    sh.cell(row, i).comment = comment
                    break
            wirte_map[material.flow_id] = index
            index = index + 1
            continue

        else:
            sh.cell(index, Consume_num).value = material.consume_num
            comment = Comment(material.zhuding_flow_id, comment_author)
            sh.cell(index, Consume_num).comment = comment
            wirte_map[material.flow_id] = index
        sh.cell(index, Brand).value = material.material_brand_id
        sh.cell(index, Factory_name).value = material.factory_name
        sh.cell(index, Flow_id).value = material.flow_id
        index = index + 1

    for i in range(0, len(material_convert)):
        material = material_convert[i]
        if material.flow_id == "" or material.flow_id == None:
            continue
        row = start_chengchanmingxi
        for j in range(4, len(manul) + 4):
            if sh_2.cell(j, Zhu_ding_pi_hao).value == None:
                row = j
            elif sh_2.cell(j, Zhu_ding_pi_hao).value == material.zhuding_flow_id:
                row = j
            else:
                row = j + 1
        if row == start_chengchanmingxi:
            sh_2.cell(start_chengchanmingxi, Pai_hao).value = material.brand_id
            sh_2.cell(start_chengchanmingxi, Ren_wu_dan_hao).value = material.task_id
            sh_2.cell(start_chengchanmingxi, Zhu_ding_pi_hao).value = material.zhuding_flow_id
            sh_2.cell(start_chengchanmingxi, Gui_ge).value = material.standard
            sh_2.cell(start_chengchanmingxi, Time_col).value = material.time
        sh_2.cell(row, Pai_hao).value = material.brand_id
        sh_2.cell(row, Ren_wu_dan_hao).value = material.task_id
        sh_2.cell(row, Zhu_ding_pi_hao).value = material.zhuding_flow_id
        sh_2.cell(row, Gui_ge).value = material.standard
        sh_2.cell(row, Time_col).value = material.time

        should_material = 0
        for j in range(11, len(Tou_liao_list) + 11):
            cell_value = sh_2.cell(3, j).value
            if cell_value == material.material_name:
                should_material = j
                break
            if cell_value == material.material_brand_id:
                should_material = j
                break
        if should_material == 0:
            continue
        cur_val = sh_2.cell(row, should_material).value
        if sh_2.cell(row, should_material).value == None:
            cur_val = 0
        a = float(material.consume_num)
        sh_2.cell(row, should_material).value = cur_val + a
    wb.save(r"C:\Users\moosk\PycharmProjects\pythonProject\Office2PDF\test\sub\result\result.xlsx")
    wb.close()


if __name__ == '__main__':
    pdf_to_excel()
    material_convert = excel_2_convert()
    print(f"material_convert {material_convert}")
    convert_2_result(material_convert)
